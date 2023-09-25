from selenium import webdriver
from selenium.webdriver.chrome.service import service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


options = webdriver.ChromeOptions()
options.add_argument('ignore-certificate-errors')


# chrome_service = webdriver.chrome.service.Service(ChromeDriverManager().install())
driver = webdriver.Chrome(options=options)

driver.get("https://10.30.0.106:8111/IRISBILLPAYMENT/UserManagement")
driver.maximize_window()
time.sleep(3)

# Load data from Excel file
excel_file = "input_data.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook["Login_creds"]

# Read username and password from Excel
username = sheet.cell(row=2, column=1).value  # Assuming username is in cell A2
password = sheet.cell(row=2, column=2).value  # Assuming password is in cell B2

Username_field = driver.find_element(By.ID, "UserName")
Username_field.send_keys(username)

password_field = driver.find_element(By.ID, "Password")
password_field.send_keys(password)

driver.find_element(By.ID, "btn_Submit").click()
# dashboard = driver.find_element(By.XPATH, "/html/body/div[10]/div[1]/div[1]").click()
# payments = driver.find_element(By.XPATH, "/html/body/div[7]/div[3]/div/div[4]/a").click()
# print("done")

# Load data from Excel file
excel_file = "input_data.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook['Elements']
# Read element names and regex locators from Excel
elements_data = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
    element_name, regex_locator = [cell.value for cell in row]
    elements_data.append((element_name, regex_locator))

# Use regex locators to find and click dashboard
element_name, regex_locator = elements_data[0]
element = driver.find_element(By.XPATH, regex_locator)
element.click()

# Use regex locators to find and click Bulk Import & Export
for element_name, regex_locator in elements_data[1:]:
    element = driver.find_element(By.XPATH, regex_locator)
    element.click()
    time.sleep(5)

    # Define actions based on element name
    if element_name == "Bulk Operations - Imports":
        button = driver.find_element(By.XPATH, "/html/body/div[10]/div[2]/div[4]/div[1]/div[1]/div/div[1]/div[2]/div[2]/div[1]/a[2]")
        button.click()

        dropdown = driver.find_element(By.CLASS_NAME, "k-input")
        dropdown.click()

        time.sleep(3)

        # Find and click the specific option you want (e.g., "-- Select --")
        option_to_select = driver.find_element(By.XPATH,"//li[text()='P0001 - CC000011 Bill Import']")
        option_to_select.click()
        time.sleep(3)

        # Locate the checkbox element by its ID (change this selector as per your HTML)
        checkbox = driver.find_element(By.ID, "IsDefaultSetting")

        # Check if the checkbox is not checked
        if not checkbox.is_selected():
            # If it's not checked, click it to check it
            checkbox.click()
            print("checkbox checked")

        # Load data from Excel file
        excel_file = "input_data.xlsx"
        workbook = load_workbook(excel_file)
        sheet = workbook['Bulk Import']


        # Read element names and regex locators from Excel
        elements_data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            element_name, regex_locator = [cell.value for cell in row]
            elements_data.append((element_name, regex_locator))

        # Use regex locators to find and click schedule type
        element_name, regex_locator = elements_data[0]
        element = driver.find_element(By.XPATH, regex_locator)
        element.click()
        time.sleep(3)

        #Use regex locators to find and click schedule type Option
        element_name, regex_locator = elements_data[1]
        element = driver.find_element(By.XPATH, f"//li[text()=\'{regex_locator}\']")
        element.click()
        print("Regex locator clicked")

        if regex_locator == "Once":
            # Use regex locators to find and click schedule type
            element_name, regex_locator = elements_data[2]
            element = driver.find_element(By.XPATH, regex_locator)
            element.click()
            print("Schedule type clicked")

            # Use regex locators to enter date & Time
            element_name, regex_locator = elements_data[3]
            element.send_keys(regex_locator)
            print("Date & Time")

            # Wait for the "Import" button to be clickable
            time.sleep(5)

            # Click Import
            import_button = driver.find_element(By.XPATH, "//input[@value='Import']")
            import_button.click()
            print("Import Clicked")


        else:

            # Wait for the "Import" button to be clickable
            time.sleep(5)
            # Click Import
            import_button = driver.find_element(By.XPATH, "//input[@value='Import']")
            import_button.click()
            print("Import Clicked")

time.sleep(5)
driver.close()
driver.quit()













        # Find and click the specific option User want
#        option_to_select = driver.find_element(By.XPATH, f"//li[text()='{regex_locator}'")
 #       option_to_select.click()















        # Locate the "schedule type" dropdown element by its selector
  #      dropdown = driver.find_element(By.XPATH, "//*[@id='schedulesettingsBox']/fieldset/div[2]/div[2]/span/span/span[1]")
   #     dropdown.click()
    #    time.sleep(3)
     #   schedule_type_dropdown = driver.find_element(By.XPATH,"//li[text()='EOD']")
      #  schedule_type_dropdown.click()





